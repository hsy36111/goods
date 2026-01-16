import streamlit as st
import pandas as pd
import numpy as np
import os
import io
import json
import openpyxl
from datetime import datetime

# ============================================================
# Streamlit 페이지 설정 (반드시 맨 처음!)
# ============================================================
st.set_page_config(page_title="최종 견적서", layout="wide")

# ============================================================
# 🔒 비밀번호 인증 시스템
# ============================================================
# 비밀번호를 변경하려면 아래 PASSWORD 변수 값을 수정하세요
PASSWORD = "goods2026"  # ← 여기서 비밀번호 변경 가능

def check_password():
    """비밀번호 확인 함수"""
    
    # 세션 상태 초기화
    if "password_correct" not in st.session_state:
        st.session_state["password_correct"] = False
    
    # 이미 인증된 경우 True 반환
    if st.session_state["password_correct"]:
        return True
    
    # 인증되지 않은 경우 로그인 화면 표시
    st.markdown("""
    <div style='text-align: center; padding: 50px 0 30px 0;'>
        <h1>🔐 굿즈 견적 시스템</h1>
        <p style='color: #666; font-size: 1.1em;'>접근하려면 비밀번호를 입력하세요</p>
    </div>
    """, unsafe_allow_html=True)
    
    # 중앙 정렬을 위한 컬럼
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        password_input = st.text_input(
            "비밀번호",
            type="password",
            placeholder="비밀번호 입력",
            key="password_input"
        )
        
        if st.button("🔓 로그인", use_container_width=True):
            if password_input == PASSWORD:
                st.session_state["password_correct"] = True
                st.success("✅ 인증 성공! 잠시 후 페이지가 로드됩니다...")
                st.rerun()
            else:
                st.error("❌ 비밀번호가 올바르지 않습니다")
        
        st.markdown("""
        <div style='text-align: center; margin-top: 50px; color: #999; font-size: 0.9em;'>
            <p>💡 비밀번호가 기억나지 않으시면 시스템 관리자에게 문의하세요</p>
        </div>
        """, unsafe_allow_html=True)
    
    return False

# 비밀번호 확인 - 인증되지 않으면 여기서 중단
if not check_password():
    st.stop()

# ============================================================
# 인증 성공 후 앱 시작
# ============================================================

st.title("🖨️ 굿즈 통합 견적 시스템")

st.markdown("<div style='margin-bottom: 10px;'></div>", unsafe_allow_html=True)
c_cust_in, _ = st.columns([0.25, 0.75])
with c_cust_in:
    customer_name = st.text_input("업체명 (Client)", placeholder="업체명을 입력하세요", key="customer_input")

def clean_text(text):
    if pd.isna(text): return ""
    text = str(text).strip() 
    if text.startswith("*"): 
        text = text[1:]
    return text

@st.cache_data
def load_data(file1, file2):
    sheets_dict = pd.read_excel(file1, sheet_name=None)
    df_price = pd.read_excel(file2)
    df_price.columns = [clean_text(col) for col in df_price.columns]
    
    if '제작 업체' in df_price.columns:
        df_price = df_price.drop(columns=['제작 업체'])
    
    try:
        wb = openpyxl.load_workbook(file2, data_only=True)
        ws = wb.active # 첫 번째 시트 사용
        
        colors_vendors = []
        for i in range(len(df_price)):
            excel_row = i + 2 
            vendor_name = np.nan
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                color = cell.fill.start_color
                
                if color and color.type == 'rgb':
                    rgb = str(color.rgb).upper()
                    if 'FFF2CC' in rgb:
                        vendor_name = '애즈랜드'
                        break 
                    elif 'FFF5F5' in rgb:
                        vendor_name = '레드프린팅'
                        break
            
            colors_vendors.append(vendor_name)
        
        df_price['제작 업체'] = colors_vendors
        df_price['제작 업체'] = df_price['제작 업체'].ffill()
            
    except Exception:
        # 색상 로드 실패 시 무시 (기본값만 사용됨)
        pass

    if '제작수량' in df_price.columns:
        df_price['제작수량'] = pd.to_numeric(df_price['제작수량'], errors='coerce')
        df_price = df_price.dropna(subset=['제작수량'])
    return sheets_dict, df_price

try:
    dict_sheets, df_input2 = load_data('input1.xlsx', 'input2.xlsx')
except Exception as e:
    st.error(f"엑셀 파일 로드 중 오류: {e}")
    st.stop()
st.markdown("""
<style>
    /* 오른쪽 컬럼 고정 (Floating Summary) */
    .summary-box {
        position: fixed;
        top: 150px;
        right: 30px;
        width: 30%;
        max-height: 80vh;
        overflow-y: auto;
        z-index: 999;
        background-color: var(--secondary-background-color);
        color: var(--text-color);
        padding: 20px;
        border: 1px solid rgba(128, 128, 128, 0.2);
        border-radius: 10px;
        box-shadow: 0px 4px 12px rgba(0,0,0,0.1);
    }
    /* 화면 너비가 좁을때는 고정 해제 (반응형) */
    @media (max-width: 1200px) {
        .summary-box {
            position: relative;
            top: 0;
            right: 0;
            width: 100%;
            box-shadow: none;
            border: none;
            padding: 0;
        }
    }
</style>
""", unsafe_allow_html=True)

c_main_L, c_main_R = st.columns([0.65, 0.35])

with c_main_L:
    col1, col2 = st.columns([1, 1])
    with col1:
        sheet_names = list(dict_sheets.keys())
        
        # Sheet 변경 시 리셋 함수
        def on_sheet_change():
            st.session_state['rows'] = [{'id': 0, 'item': None, 'spec': None, 'margin': None, 'designs': []}]
            st.session_state['next_id'] = 1
        
        selected_sheet = st.selectbox(
            "굿즈 종류 (Sheet)", 
            sheet_names, 
            key="selected_sheet",
            on_change=on_sheet_change
        )

    with col2:
        base_qty = st.number_input("전체 제작 수량 (EA)", min_value=10, value=100, step=10, key="base_qty")
    
    st.write("") 
    # 컨테이너로 감싸서 구분
    with st.container():
        # 2개의 큰 컬럼으로 분할 (급행 / 할인)
        c_grp1, c_grp2 = st.columns(2)
        
        # 1. 급행 그룹
        with c_grp1:
            sub_c1, sub_c2 = st.columns([0.4, 0.6])
            with sub_c1:
                is_express = st.checkbox("🚀 급행 적용", value=False, help="체크 시 전체 단가에 할증이 적용됩니다.", key="is_express")
            with sub_c2:
                if is_express:
                    express_rate = st.number_input("할증률", value=1.2, step=0.1, format="%.2f", label_visibility="collapsed", key="express_rate")
                else:
                    express_rate = 1.0
        global_discount_amt = 0
        with c_grp2:
            sub_c3, sub_c4 = st.columns([0.4, 0.6])
            with sub_c3:
                is_global_discount = st.checkbox("💸 할인 적용", value=False, help="체크 시 전체 금액에서 차감됩니다.", key="is_global_discount")
            with sub_c4:
                if is_global_discount:
                    global_discount_amt = st.number_input("할인액(원)", value=50000, min_value=0, step=1000, label_visibility="collapsed", key="global_discount_amt")

    df_options = dict_sheets[selected_sheet]
    df_options['품명'] = df_options['품명'].apply(clean_text)

    st.markdown("---")
    st.markdown("### 🧾 견적 구성")
    h1, h2, h3, h4 = st.columns([1.8, 1.8, 1.8, 0.4])
    h1.markdown("**품명**")
    h2.markdown("**규격**")
    h3.markdown("**금액 / 업체**")
    st.divider()

    # =========================================================
    # 3. 데이터 및 로직 (Rows)
    # =========================================================


if 'rows' not in st.session_state:
    st.session_state['rows'] = [{'id': 0, 'item': None, 'spec': None, 'margin': None, 'designs': []}]
if 'next_id' not in st.session_state:
    st.session_state['next_id'] = 1

def get_vendor_badge_html(vendor_name):
    if pd.isna(vendor_name) or str(vendor_name).strip() == "":
        return ""
    v_name = str(vendor_name).strip()
    
    bg_color = "#e0e0e0"
    text_color = "#000000"
    if "애즈랜드" in v_name: bg_color = "#FFF2CC"
    elif "레드프린팅" in v_name: bg_color = "#FFF5F5"
    
    style = f"background-color:{bg_color}; color:{text_color}; padding:2px 8px; border-radius:10px; font-weight:bold; font-size:0.8em; margin-top:2px; display:inline-block;"
    return f"<span style='{style}'>{v_name}</span>"


def calculate_single_design_cost(design_qty, df_input2, spec_clean, margin_rate, express_rate=1.0):
    """
    단일 도안에 대한 가격 계산 (반올림 로직 포함)
    """
    result = {'price': 0, 'vendor': None, 'note': '', 'success': False, 'matched_qty': 0}
    
    qty_col = '제작수량'
    if qty_col not in df_input2.columns:
        result['note'] = "❌단가표(수량) 없음"
        return result

    # 1. 수량 반올림 (Table Lookup)
    # 25 -> 30, 50 -> 50
    avail_qtys = df_input2[qty_col].sort_values()
    matched_qty = avail_qtys[avail_qtys >= design_qty].min()
    
    # 매칭되는 수량이 없으면(최대치 초과 등) 가장 큰 값? 혹은 에러? 
    # 일단 가장 큰 값으로 fallback 또는 max
    if pd.isna(matched_qty): 
        matched_qty = avail_qtys.max()
    
    result['matched_qty'] = matched_qty

    matched_row = df_input2[df_input2[qty_col] == matched_qty]
    
    if matched_row.empty:
        result['note'] = f"❌데이터 없음({matched_qty})"
        return result

    # 업체 확인
    if '제작 업체' in matched_row.columns:
        val_vendor = matched_row['제작 업체'].values[0]
        if pd.notna(val_vendor):
            result['vendor'] = val_vendor

    # 가격 확인
    if spec_clean in matched_row.columns:
        base_cost = matched_row[spec_clean].values[0]
        # [UPDATE] 급행 할증 / 마진 적용
        final_cost = base_cost * margin_rate * express_rate
        result['price'] = final_cost
        result['success'] = True
    else:
        result['note'] = f"❌규격 매칭 실패"

    return result

def calculate_cost(row_data, base_qty, df_input2, override_margin=None, express_rate=1.0):
    scope = str(row_data['제작수량']).strip()
    unit_base = row_data['단가/*수량']
    ref_val = row_data['참조 값']
    vendor = row_data['제작 업체'] 
    spec_clean = clean_text(row_data['규격'])
    
    if pd.isna(unit_base): unit_base = 0
    if pd.isna(ref_val): ref_val = 0

    result = {
        'price': 0,
        'unit_price': 0,
        'note': '',
        'vendor': vendor,
        'success': True
    }

    try:
        # CASE 1: UNIT_QTY -> Table Lookup
        if scope == 'UNIT_QTY':
            # 마진 배율 적용 (Override 있으면 우선 사용)
            margin_rate = override_margin if override_margin is not None else ref_val
            
            qty_col = '제작수량'
            if qty_col in df_input2.columns:
                avail_qtys = df_input2[qty_col].sort_values()
                matched_qty = avail_qtys[avail_qtys >= base_qty].min()
                if pd.isna(matched_qty): matched_qty = avail_qtys.max()
                
                matched_row = df_input2[df_input2[qty_col] == matched_qty]
                
                # [Fix] 색상으로 찾은 업체 우선 적용
                # df_input2에 이미 색상 기반 '제작 업체'가 ffill 등으로 채워져 있음
                if '제작 업체' in matched_row.columns:
                    val_vendor = matched_row['제작 업체'].values[0]
                    if pd.notna(val_vendor):
                        result['vendor'] = val_vendor

                if spec_clean in matched_row.columns:
                    base_cost = matched_row[spec_clean].values[0]
                    final_cost = base_cost * margin_rate * express_rate
                    
                    result['price'] = final_cost
                    result['unit_price'] = final_cost
                    qty_disp = int(matched_qty) if matched_qty % 1 == 0 else matched_qty
                    margin_disp = round(margin_rate, 2)
                    
                    note_parts = [f"{qty_disp}개 구간", f"마진 {margin_disp}x"]
                    if express_rate > 1.0:
                        note_parts.append(f"🚀급행 {express_rate}x")
                    
                    result['note'] = " (".join(note_parts) + ")" if len(note_parts) > 1 else note_parts[0]
                else:
                    result['note'] = f"❌표 매칭 실패({spec_clean})"
                    result['success'] = False
            else:
                result['note'] = "❌단가표(수량) 없음"
                result['success'] = False

        elif scope == 'GLOBAL_QTY':
            unit_price = (unit_base + ref_val) * express_rate
            result['price'] = unit_price * base_qty
            result['unit_price'] = unit_price

        else:
            unit_price = (unit_base + ref_val) * express_rate
            result['price'] = unit_price
            result['unit_price'] = unit_price
            result['note'] = "고정비" + (f" (🚀{express_rate}x)" if express_rate > 1.0 else "")

    except Exception as e:
        result['note'] = f"Error: {str(e)}"
        result['success'] = False
        
    return result

from openpyxl.styles import Font, Color

def generate_excel_from_template(data_list, total, vat, grand_total, sheet_name, base_qty, customer_name="", discount_amt=0):
    try:
        # 템플릿 로드 (UI.xlsx가 같은 폴더에 있다고 가정)
        template_path = 'UI.xlsx'
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active # 첫 번째 시트 사용
        
        # [Update] 1. 제목 (C3) - 괄호 제거
        ws['C3'] = f"{sheet_name} 견적서"
        
        # [Update] 문서번호 (C6) & 견적일 (C7) & 업체명 (B10)
        today = datetime.now()
        ws['C6'] = today.strftime("%y%m%d") + "_"
        ws['C7'] = today.strftime("%Y-%m-%d")
        # [Update] 업체명 + 담당자님 표기
        ws['B10'] = f"{customer_name} 담당자님" if customer_name else ""
        
        # 2. 전체 수량 (D14)
        ws['D14'] = base_qty
        
        # 3. 품목 리스트 (17행 ~ 26행, 최대 10개)
        # data_list에는 순수 품목만 있음 (할인 제외)
        start_row = 17
        max_rows = 10
        
        for i, item in enumerate(data_list):
            if i >= max_rows:
                break # 10개까지만 입력 가능
                
            row_idx = start_row + i
            
            # 품명 (B열)
            ws[f'B{row_idx}'] = item.get('품명', '')
            # 규격 (C열)
            ws[f'C{row_idx}'] = item.get('규격', '')
            # 금액 (F열)
            ws[f'F{row_idx}'] = item.get('금액', 0)
            # 비고/도안요약 (I열)
            ws[f'I{row_idx}'] = item.get('단가_비고', '') # 도안 요약 정보가 여기 들어있음
            # 제작 업체 (J열)
            ws[f'J{row_idx}'] = item.get('업체', '')
            
        # [Update] 할인 항목 별도 기입 (B27~B29 영역)
        # 유저 요청: B27, B28, B29 쪽에 대입. 
        # B27에 기입 (Listが終わった直後)
        if discount_amt > 0:
            target_row = 27
            ws[f'B{target_row}'] = "할인"
            # 글자 빨간색 처리
            ws[f'B{target_row}'].font = Font(color="FF0000") 
            
            ws[f'F{target_row}'] = -discount_amt # 차감액 (음수 표시)
            ws[f'I{target_row}'] = "Global Discount"

        # 4. 최종 금액
        # E29와 E30이 셀병합 되어 있음 -> E29에 입력
        ws['E29'] = total
        
        # [Update] User Request: E30 and F30 are merged -> Grand Total here
        ws['E30'] = grand_total 
        
        # 5. 개당 단가
        # H40: VAT 미포함, H41: VAT 포함
        if base_qty > 0:
            ws['H40'] = total / base_qty
            ws['H41'] = grand_total / base_qty
            
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"템플릿 엑셀 생성 중 오류: {e}")
        return None

# [Layout] Left Column Content (Inputs & Rows)
with c_main_L:
    # -----------------------------------------------------------
    # Rows Logic (Hybrid Input)
    # -----------------------------------------------------------

    remove_ids = []
    raw_items = df_options['품명'].unique()
    item_options = [x for x in raw_items if pd.notna(x) and str(x).strip() != ""]


    export_data = []
    total_supply_price = 0
    total_gross_price = 0      # 총 공급가액 (할인 전)
    total_discount_amount = 0  # 총 할인 금액

    # [Ref Data] 수기 입력 대상 불러오기


    for i, row in enumerate(st.session_state['rows']):
        row_id = row['id']
        
        # 레이아웃: 품명(2) | 규격(2) | 금액(2) | 삭제(0.4)
        c1, c2, c3, c4 = st.columns([1.8, 1.8, 1.8, 0.4])
        
        # State Keys for syncing
        # 우리는 row['item'], row['spec']을 text_input 값으로 유지.
        # Selectbox는 단순 Helper 역할.
        
        with c1:
            # -----------------------------------------------------------
            # 1. 품명 (Item) - Smart Dual Input
            # -----------------------------------------------------------
            row_id = row['id']
            current_item = row['item'] if row['item'] else ""
            
            # 품명이 목록에 있는지 확인
            is_in_options = current_item in item_options
            
            # [Helper] 검색/선택용 Dropdown
            def on_item_help_change(r_id=row_id):
                val = st.session_state.get(f"isel_{r_id}")
                
                target_idx = -1
                for idx, r in enumerate(st.session_state['rows']):
                    if r['id'] == r_id:
                        target_idx = idx
                        break
                
                if target_idx != -1:
                    # 값이 있으면 선택된 값으로, 없으면(X버튼) 기존 값 유지 or 초기화?
                    # UX: X버튼 누르면 수기 입력을 위해 비워주는게 좋음
                    new_val = val if val else "" 
                    st.session_state['rows'][target_idx]['item'] = new_val
                    st.session_state['rows'][target_idx]['spec'] = ""

            # 선택된 상태면 index 설정, 아니면 None (수기모드)
            sel_index = item_options.index(current_item) if is_in_options else None
            
            st.selectbox(
                "품명 선택",
                options=item_options,
                index=sel_index,
                key=f"isel_{row_id}",
                placeholder="품명 검색 (선택하세요)",
                label_visibility="collapsed",
                on_change=on_item_help_change,
                kwargs={'r_id': row_id}
            )

            # [Main] 실제 입력값 (Text Input)
            # 목록에 없는 값(수기)이거나 빈칸일 때만 노출
            # 목록에서 선택된 상태면 숨김 (요청사항: "선택되었으면 하단 칸 없애줘")
            if not is_in_options:
                final_item = st.text_input(
                    "품명 직접 입력", 
                    value=current_item, 
                    key=f"itxt_{row_id}", 
                    placeholder="품명 직접 입력 (목록에 없을 시)",
                    label_visibility="collapsed"
                )
            else:
                final_item = current_item
            
            # Update State
            if final_item != st.session_state['rows'][i]['item']:
                 st.session_state['rows'][i]['item'] = final_item
                 st.session_state['rows'][i]['spec'] = ""

        with c2:
            # -----------------------------------------------------------
            # 2. 규격 (Spec) - Smart Dual Input with 2-Tier Dropdown
            # -----------------------------------------------------------
            final_spec = ""
            
            if final_item:
                # 규격 목록 가져오기
                spec_opts = []
                if final_item in item_options:
                    item_df = df_options[df_options['품명'] == final_item]
                    spec_opts = item_df['규격'].tolist()
                
                current_spec = row['spec'] if row['spec'] else ""
                
                # [NEW] * 체크: 규격 옵션 중 하나라도 *로 시작하면 2단계 드롭다운 활성화
                has_asterisk = any(str(s).startswith('*') for s in spec_opts if pd.notna(s))
                
                if has_asterisk:
                    # === 2단계 드롭다운 시스템 ===
                    tier_map = {}  # {1뎁스: [2뎁스 옵션들]}
                    tier1_order = []  # 등장 순서 보존
                    
                    for spec_raw in spec_opts:
                        if pd.isna(spec_raw):
                            continue
                        spec_clean = str(spec_raw).lstrip('*')  # * 제거
                        
                        if '|' in spec_clean:
                            parts = spec_clean.split('|', 1)
                            tier1 = parts[0].strip()
                            tier2 = parts[1].strip()
                            
                            if tier1 not in tier_map:
                                tier_map[tier1] = []
                                tier1_order.append(tier1)  # 순서 기록
                            if tier2 not in tier_map[tier1]:
                                tier_map[tier1].append(tier2)
                        else:
                            # | 없는 경우 (예: 스티커)
                            tier1 = spec_clean.strip()
                            if tier1 not in tier_map:
                                tier_map[tier1] = []
                                tier1_order.append(tier1)  # 순서 기록
                    
                    # 현재 선택값 파싱
                    current_tier1 = ""
                    current_tier2 = ""
                    
                    if current_spec:
                        spec_for_parse = current_spec.lstrip('*')
                        if '|' in spec_for_parse:
                            parts = spec_for_parse.split('|', 1)
                            current_tier1 = parts[0].strip()
                            current_tier2 = parts[1].strip()
                        else:
                            current_tier1 = spec_for_parse.strip()
                    
                    # 1뎁스 선택 (등장 순서 유지)
                    tier1_options = tier1_order  # sorted 제거
                    tier1_index = tier1_options.index(current_tier1) if current_tier1 in tier1_options else 0
                    
                    selected_tier1 = st.selectbox(
                        "1뎁스",
                        options=tier1_options,
                        index=tier1_index,
                        key=f"tier1_{row_id}",
                        label_visibility="collapsed"
                    )
                    
                    # 2뎁스 선택
                    tier2_options = tier_map.get(selected_tier1, [])
                    
                    if tier2_options:
                        tier2_index = tier2_options.index(current_tier2) if current_tier2 in tier2_options else 0
                        
                        selected_tier2 = st.selectbox(
                            "2뎁스",
                            options=tier2_options,
                            index=tier2_index,
                            key=f"tier2_{row_id}",
                            label_visibility="collapsed"
                        )
                        final_spec = f"{selected_tier1}|{selected_tier2}"
                    else:
                        st.text_input(
                            "2뎁스 (없음)",
                            value="",
                            disabled=True,
                            key=f"tier2_disabled_{row_id}",
                            label_visibility="collapsed"
                        )
                        final_spec = selected_tier1
                    
                    if final_spec != current_spec:
                        st.session_state['rows'][i]['spec'] = final_spec
                else:
                    # === 기존 단일 드롭다운 ===
                    is_spec_in_opts = current_spec in spec_opts
                
                    # [Helper] 규격 선택 도우미
                    def on_spec_help_change(r_id=row_id):
                        val = st.session_state.get(f"ssel_{r_id}")
                        
                        target_idx = -1
                        for idx, r in enumerate(st.session_state['rows']):
                            if r['id'] == r_id:
                                target_idx = idx
                                break

                        if target_idx != -1:
                             new_val = val if val else ""
                             st.session_state['rows'][target_idx]['spec'] = new_val

                    s_index = spec_opts.index(current_spec) if is_spec_in_opts else None

                    st.selectbox(
                        "규격 선택",
                        options=spec_opts,
                        index=s_index,
                        key=f"ssel_{row_id}",
                        placeholder="규격 선택",
                        label_visibility="collapsed",
                        on_change=on_spec_help_change,
                        kwargs={'r_id': row_id},
                        disabled=(len(spec_opts) == 0)
                    )

                    #  [Main] 실제 규격 입력값
                    # 규격이 선택되었으면 숨김
                    if not is_spec_in_opts:
                        final_spec = st.text_input(
                            "규격 직접 입력", 
                            value=current_spec, 
                            key=f"stxt_{row_id}", 
                            placeholder="규격 직접 입력",
                            label_visibility="collapsed"
                        )
                    else:
                        final_spec = current_spec
                    
                    if final_spec != st.session_state['rows'][i]['spec']:
                        st.session_state['rows'][i]['spec'] = final_spec
            else:
                st.write("-")

        # -----------------------------------------------------------
        # 3. 금액 계산 (Price & Vendor)
        # -----------------------------------------------------------
        with c3:
            if final_item and final_spec:
                try:
                    # [NEW] X 체크: 규격이 |X 로 끝나면 0원 처리
                    is_x_spec = False
                    if '|' in final_spec:
                        parts = final_spec.split('|')
                        if len(parts) >= 2 and parts[-1].strip().upper() == 'X':
                            is_x_spec = True
                    
                    if is_x_spec:
                        # X 선택 시: 0원 표시, 도안 분배 제외
                        st.markdown("**0 원**")
                        st.caption("❌ 제외 항목")
                        
                        export_data.append({
                            '품명': final_item, 
                            '규격': final_spec, 
                            '단가_비고': '제외 항목 (X)', 
                            '금액': 0, 
                            '업체': '-'
                        })
                        # total_supply_price에 더하지 않음 (0원)
                        
                    else:
                        # 기존 로직: DB 매칭 시도
                        matched_row = pd.DataFrame()
                        
                        # [NEW] * 제거 (Input2 조회 시)
                        spec_for_lookup = final_spec.lstrip('*')
                        
                        # DB에 존재하는 품명/규격인지 확인
                        if final_item in item_options:
                             item_df = df_options[df_options['품명'] == final_item]
                             # 규격 매칭 (* 제거된 값으로 매칭)
                             matched_row = item_df[item_df['규격'].str.lstrip('*') == spec_for_lookup]
                        
                        # ---------------------------
                        # A. DB 매칭 성공
                        # ---------------------------
                        if not matched_row.empty:
                            data_row = matched_row.iloc[0]
                            
                            price_val = 0
                            vendor_val = data_row.get('제작 업체', np.nan)
                            unit_val_display = ""
                            
                            # [Logic] UNIT_QTY 일 경우 마진 수정 기능
                            scope_val = str(data_row.get('제작수량', '')).strip()
                            ref_margin_val = data_row.get('참조 값', 1.0)
                            if pd.isna(ref_margin_val): ref_margin_val = 1.0
                            
                            override_margin_val = None
                            
                            if scope_val == 'UNIT_QTY':
                                # State에 마진값이 없으면 초기화
                                if row.get('margin') is None:
                                     st.session_state['rows'][i]['margin'] = float(ref_margin_val)
                                
                                current_margin = st.session_state['rows'][i]['margin']
                                
                                c_p1, c_p2 = st.columns([0.4, 0.6])
                                with c_p1:
                                    st.caption("마진배율")
                                with c_p2:
                                    new_margin = st.number_input(
                                        "마진", 
                                        value=current_margin, 
                                        step=0.1, 
                                        format="%.1f",
                                        key=f"margin_{row_id}", 
                                        label_visibility="collapsed"
                                    )
                                
                                # 값 업데이트
                                if new_margin != current_margin:
                                    st.session_state['rows'][i]['margin'] = new_margin
                                
                                override_margin_val = new_margin

                            # 특수 로직 (택배/퀵)
                            if "택배" in final_spec:
                                c_in1, c_in2 = st.columns([1,1])
                                with c_in1: box = st.number_input("박스", 1, value=1, key=f"bx_{row_id}", label_visibility="collapsed")
                                with c_in2: price_val = box*3000; st.write(f"**{price_val:,}**")
                                vendor_val="택배"; unit_val_display=f"3,000 x {box}"
                                
                            elif "퀵" in final_spec:
                                quick = st.number_input("퀵비", 0, step=1000, key=f"qk_{row_id}", label_visibility="collapsed")
                                price_val = quick
                                vendor_val="퀵/용달"; unit_val_display="실비"
                                
                            else:
                                # 일반 계산
                                calc_res = calculate_cost(data_row, base_qty, df_input2, override_margin=override_margin_val, express_rate=express_rate)
                                price_val = calc_res['price']
                                if pd.notna(calc_res['vendor']): vendor_val = calc_res['vendor']
                                
                                badge = get_vendor_badge_html(vendor_val)
                                # [Update] 업체명 줄바꿈 처리 (<br> 추가)
                                st.markdown(f"<div style='line-height:1.4;'><b>{int(price_val):,} 원</b><br>{badge}</div>", unsafe_allow_html=True)
                                if calc_res['note']: 
                                    st.caption(calc_res['note']); unit_val_display = calc_res['note']
                                    
                            export_data.append({'품명': final_item, '규격': final_spec, '단가_비고': unit_val_display, '금액': price_val, '업체': vendor_val})
                            total_supply_price += price_val
                            total_gross_price += price_val # 총 공급가액 (양수) 누적


                            # [Feature] 내지 및 스티커 제작 - 도안 분배 기능
                            # 조건: 품명에 "내지 및 스티커 제작" 또는 "컨러군번줄" 포함 시
                            is_design_split = "내지 및 스티커 제작" in final_item
                            is_color_split = "컨러군번줄" in final_item
                            
                            if is_design_split or is_color_split:
                                st.markdown("---")
                                heading = "🎨 **색상 분배 (Multi-Color)**" if is_color_split else "🎨 **도안 분배 (Multi-Design)**"
                                st.caption(heading)
                                
                                # 도안 데이터 초기화 (없으면 기본값)
                                if 'designs' not in row or not row['designs']:
                                 # 초기: 1개 = 전체수량
                                 # [Update] Multi-Spec일 경우 spec 필드도 관리해야 함 (None)
                                  st.session_state['rows'][i]['designs'] = [{'id': 0, 'qty': base_qty, 'spec': None}]
                                
                                designs = st.session_state['rows'][i]['designs']
                                
                                # 추가 버튼
                                btn_label = "➕ 색상 추가" if is_color_split else "➕ 도안 추가"
                                if st.button(btn_label, key=f"add_ds_{row_id}"):
                                    new_d_id = max([d.get('id', 0) for d in designs]) + 1 if designs else 0
                                    designs.append({'id': new_d_id, 'qty': 0, 'spec': None})
                                    st.rerun()

                                total_design_price = 0
                                design_breakdown = []
                                design_results = [] # (design, result_dict) list
                            
                            # 도안별 입력 및 계산 반복
                                for d_idx, design in enumerate(designs):
                                    d_id = design.get('id', d_idx)
                                
                                    # Layout
                                    if is_color_split:
                                        # [Multi-Color Layout] 색상(크게) | 수량(절반) | 삭제
                                        c_d1, c_d1_5, c_d3 = st.columns([3.0, 1.0, 0.3])
                                        c_d2 = None  # Price info will be shown below
                                    else:
                                        # [Multi-Design Layout] Qty | Del
                                        c_d1, c_d3 = st.columns([3.0, 0.3])
                                        c_d1_5 = None
                                        c_d2 = None
                                
                                    # 1. Spec/Color Selection (Only for Color Split)
                                    target_spec = final_spec
                                
                                    if is_color_split and c_d1_5:
                                        with c_d1:
                                            # 색상 리스트 가져오기
                                            sub_spec_opts = []
                                            if final_item in item_options:
                                                tmp_df = df_options[df_options['품명'] == final_item]
                                                sub_spec_opts = tmp_df['규격'].tolist()
                                        
                                            # 현재 선택된 spec (없으면 첫번째 or parent spec)
                                            cur_sub_spec = design.get('spec')
                                            # 만약 cur_sub_spec이 옵션에 없으면 기본값 설정
                                            idx_sel = 0
                                            if cur_sub_spec in sub_spec_opts:
                                                idx_sel = sub_spec_opts.index(cur_sub_spec)
                                            else:
                                                # 초기값: final_spec이 옵션에 있으면 그걸로, 아니면 첫번째
                                                if final_spec in sub_spec_opts:
                                                    idx_sel = sub_spec_opts.index(final_spec)
                                        
                                            selected_sub_spec = st.selectbox(
                                                f"색상 {d_idx+1}", 
                                                sub_spec_opts, 
                                                index=idx_sel,
                                                key=f"dsp_{row_id}_{d_id}", 
                                                label_visibility="collapsed"
                                            )
                                        
                                            # State Update
                                            if selected_sub_spec != cur_sub_spec:
                                                st.session_state['rows'][i]['designs'][d_idx]['spec'] = selected_sub_spec
                                                # Need to ensure next calculations use this
                                        
                                            target_spec = selected_sub_spec
                                
                                    # 2. Quantity Input
                                    # Qty Input Location depends on layout
                                    qty_col = c_d1_5 if is_color_split else c_d1
                                    label_txt = f"색상 {d_idx+1} 수량" if is_color_split else f"도안 {d_idx+1} 수량"
                                
                                    with qty_col:
                                        d_qty = st.number_input(
                                            label_txt, 
                                            min_value=0, 
                                            value=int(design.get('qty', 0)),
                                            step=1,
                                            key=f"d_qty_{row_id}_{d_id}",
                                            label_visibility="collapsed"
                                        )
                                        if d_qty != design.get('qty'):
                                            st.session_state['rows'][i]['designs'][d_idx]['qty'] = d_qty
                                
                                    # 개별 가격 계산 (하단에 표시할 예정)
                                    applied_margin = override_margin_val if override_margin_val is not None else ref_margin_val
                                
                                    if is_color_split:
                                        # [Update] 컨러군번줄 등 Multi-Color는 Input1의 단가(E열) * 수량 으로 단순 계산
                                        d_res = {'price':0, 'vendor':np.nan, 'success':False, 'matched_qty':d_qty, 'note':''}
                                    
                                        try:
                                            cat_df = df_options[(df_options['품명'] == final_item) & (df_options['규격'] == target_spec)]
                                        
                                            if not cat_df.empty:
                                                unit_base_val = cat_df.iloc[0]['단가/*수량']
                                                if pd.isna(unit_base_val): unit_base_val = 0
                                            
                                                calc_price = unit_base_val * d_qty * express_rate
                                            
                                                d_res['price'] = calc_price
                                                d_res['vendor'] = cat_df.iloc[0]['제작 업체'] if '제작 업체' in cat_df.columns else None
                                                d_res['success'] = True
                                            else:
                                                d_res['note'] = "❌Input1 매칭 실패"
                                            
                                        except Exception as e:
                                            d_res['note'] = str(e)

                                    else:
                                        # Existing Multi-Design Logic (Input2 Lookup)
                                        d_res = calculate_single_design_cost(d_qty, df_input2, clean_text(target_spec), applied_margin, express_rate=express_rate)
                                
                                    # 가격 정보 표시 (입력필드 하단에)
                                    if d_res['success']:
                                        p_display = int(d_res['price'])
                                        v_display = d_res['vendor'] if d_res['vendor'] else ""
                                        matched_q = int(d_res['matched_qty'])
                                    
                                        badge_html = get_vendor_badge_html(v_display)
                                        st.markdown(f"➜ **{p_display:,}원** ({matched_q}개) {badge_html}", unsafe_allow_html=True)
                                    
                                        total_design_price += d_res['price']
                                    
                                        summ_txt = f"{target_spec}({d_qty}ea)" if is_color_split else f"도안{d_idx+1}({d_qty}ea)"
                                        design_breakdown.append(f"{summ_txt}: {p_display:,}")
                                    
                                        d_res['vendor'] = v_display 
                                        design_results.append((design, d_res)) 
                                    else:
                                        # [Update] 에러 메시지 미표시 (User Request)
                                        pass
                                        design_results.append((design, {'vendor': None}))
                                
                                    with c_d3:
                                        # 삭제 버튼
                                        if len(designs) > 1:
                                            if st.button("x", key=f"del_ds_{row_id}_{d_id}"):
                                                st.session_state['rows'][i]['designs'].pop(d_idx)
                                                st.rerun()

                                # 최종 합계 표시 및 메인 가격 덮어쓰기
                                if not is_color_split:
                                    st.markdown(f"👉 **도안 합계: {int(total_design_price):,} 원**")
                            
                                # [Important] 메인 로직의 결과값을 이 도안 합계로 대체해야 함
                                # 위에서 이미 export_data.append(...) 하고 total_supply_price += ... 했음.
                                # 이를 취소하고 덮어써야 함.
                            
                                # 1. 방금 더한 price_val 차감
                                total_supply_price -= price_val
                                total_gross_price -= price_val
                            
                                # 2. 새로운 값 더하기
                                price_val = total_design_price
                                total_supply_price += price_val
                                total_gross_price += price_val
                            
                                # 3. Vendor Aggregation (모든 도안의 업체 수집)
                                # [Fix] "sequence item 0: expected str instance" 오류 해결
                                # 업체명이 float(NaN)이나 다른 타입일 경우 문자열로 변환하여 처리
                                def safe_str(v):
                                    if pd.isna(v) or str(v).strip() == "": return None
                                    return str(v).strip()

                                vendors_raw = [safe_str(d_res.get('vendor')) for d, d_res in design_results]
                                # None 제거 및 중복 제거
                                unique_vendors = sorted(list(set([v for v in vendors_raw if v is not None])))
                            
                                if unique_vendors:
                                     vendor_val = " / ".join(unique_vendors)
                                else:
                                     vendor_val = ""

                                # 4. Export Data 수정 (마지막에 추가된 항목 pop 후 다시 append)
                                if export_data:
                                    export_data.pop()
                                    note_str = " / ".join(design_breakdown)
                                    if express_rate > 1.0:
                                        note_str += f" (🚀급행 {express_rate:.1f}x)"
                                
                                    export_data.append({
                                        '품명': final_item, 
                                        '규격': final_spec, 
                                        '단가_비고': f"[도안분배] {note_str}", 
                                        '금액': price_val, 
                                        '업체': vendor_val 
                                    })
                            
                            # ---------------------------
                            # B. DB 매칭 실패 (수기 입력)
                            # ---------------------------
                        else:
                            # 사용자에게 단가 입력 요구
                            c_m1, c_m2 = st.columns([0.4, 0.6])
                            with c_m1:
                                st.caption("단가(별도):")
                            with c_m2:
                                 # 기존에 입력된 값이 있다면 유지? -> st.session_state에 따로 저장 안하면 0 초기화됨.
                                 # key가 row_id에 종속되므로 유지됨.
                                 manual_unit_price = st.number_input("단가", min_value=0, step=100, key=f"m_price_{row_id}", label_visibility="collapsed")
                            
                            price_val = manual_unit_price * base_qty
                            vendor_val = "수기입력"
                            unit_val_display = f"@{manual_unit_price:,}"
                            
                            st.markdown(f"합계: **{int(price_val):,} 원**")
                            
                            export_data.append({'품명': final_item, '규격': final_spec, '단가_비고': unit_val_display, '금액': price_val, '업체': vendor_val})
                            total_supply_price += price_val
                            total_gross_price += price_val # 총 공급가액 누적
                        
                except Exception as e:
                    st.error(f"오류: {e}")
            else:
                st.write("-")

        with c4:
            st.markdown("<div style='margin-top: 5px;'></div>", unsafe_allow_html=True)
            if st.button("🗑️", key=f"del_{row_id}"): remove_ids.append(row_id)
        
        # [NEW] SPC 규격 추가 버튼
        if final_item and final_spec:
            try:
                # DB에서 해당 품목/규격 찾기
                check_row = pd.DataFrame()
                if final_item in item_options:
                    item_df_check = df_options[df_options['품명'] == final_item]
                    spec_check = final_spec.lstrip('*')
                    check_row = item_df_check[item_df_check['규격'].str.lstrip('*') == spec_check]
                
                if not check_row.empty:
                    row_data = check_row.iloc[0]
                    
                    # G열 값 가져오기 (컨럼 인덱스 6, 0-based)
                    spc_value = None
                    if len(check_row.columns) > 6:
                        spc_value = row_data.iloc[6] if len(row_data) > 6 else None
                    
                    # SPC 확인
                    is_spc = str(spc_value).strip().upper() == "SPC" if pd.notna(spc_value) else False
                    
                    if is_spc:
                        st.markdown("<div style='margin-top: 10px;'></div>", unsafe_allow_html=True)
                        if st.button("➕ 규격 추가 (동일 품명)", key=f"add_spec_{row_id}", use_container_width=True):
                            # 같은 품명으로 새 행 추가
                            new_row = {
                                'id': st.session_state['next_id'],
                                'item': final_item,
                                'spec': None,
                                'margin': None,
                                'designs': []
                            }
                            st.session_state['rows'].append(new_row)
                            st.session_state['next_id'] += 1
                            st.rerun()
            except:
                pass
        
        st.divider()

    if remove_ids:
        st.session_state['rows'] = [r for r in st.session_state['rows'] if r['id'] not in remove_ids]
        st.rerun()

    if st.button("➕ 품목 추가하기"):
        st.session_state['rows'].append({'id': st.session_state['next_id'], 'item': None, 'spec': None, 'margin': None, 'designs': []})
        st.session_state['next_id'] += 1
        st.rerun()

# [Layout] Right Column Content (Summary)
with c_main_R:
    # [Update] 요약창은 항상 표시 (Fixed Layout 유지를 위해)
    # [Update] export_data에 '전체 할인' 항목 추가하지 않음 (엑셀 함수에 별도 전달)
    if global_discount_amt > 0:
        total_discount_amount += global_discount_amt
        total_supply_price -= global_discount_amt
        # export_data.append({
        #     '품명': '전체 할인', 
        #     '규격': 'Global Discount', 
        #     '단가_비고': '할인 적용', 
        #     '금액': -global_discount_amt, 
        #     '업체': '할인'
        # })

    vat = total_supply_price * 0.1
    grand_total = total_supply_price + vat
    
    # 할인율 계산
    discount_rate_str = "-"
    if total_gross_price > 0:
        rate = (total_discount_amount / total_gross_price) * 100
        discount_rate_str = f"{rate:.2f} %"
        
    per_unit_supply = total_supply_price / base_qty if base_qty > 0 else 0
    per_unit_grand = grand_total / base_qty if base_qty > 0 else 0
    
    # 엑셀 파일 생성 (미리 생성하여 링크로 제공)
    excel_b64 = None
    if total_supply_price != 0: # 0원이 아닐때만 다운로드 생성
        try:
            # [Update] discount_amt 인자 전달
            excel_data = generate_excel_from_template(export_data, total_supply_price, vat, grand_total, selected_sheet, base_qty, customer_name=customer_name, discount_amt=global_discount_amt)
            if excel_data:
                import base64
                excel_b64 = base64.b64encode(excel_data).decode()
        except Exception as e:
            st.error(f"엑셀 생성 오류: {e}")

    # Download Link HTML
    # [Fix] HTML이 코드로 인식되지 않도록 공백 제거 (Dedented)
    # [Update] 파일명 형식 변경: YYMMDD_업체명 견적서(굿즈종류).xlsx
    today_str_file = datetime.now().strftime("%y%m%d")
    cust_str_file = customer_name.strip() if customer_name else "업체미지정"
    file_name = f"{today_str_file}_{cust_str_file} 견적서({selected_sheet}).xlsx"
    
    download_html = ""
    if excel_b64:
        download_html = f'''<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_b64}" download="{file_name}" style="text-decoration:none;"><div style="background-color: #4CAF50; color: white; padding: 10px 15px; text-align: center; border-radius: 5px; font-weight: bold; margin-top: 15px; cursor: pointer;">📥 엑셀 견적서 다운로드</div></a>'''
    elif total_supply_price == 0:
        download_html = "<div style='margin-top:15px; text-align:center; color:#aaa;'>견적 내용이 없습니다</div>"

    # Floating Box 전체를 하나의 HTML로 묶어 Rendering
    st.markdown(f"""<div class="summary-box">
<div style='text-align: right;'>
<div style='font-size: 1.0rem; margin-bottom: 5px; opacity: 0.8;'>할인 전 금액 : {int(total_gross_price):,} 원</div>
<div style='font-size: 1.0rem; margin-bottom: 5px; color: #448AFF;'>총 할인액 : -{int(total_discount_amount):,} 원 (할인율 : {discount_rate_str})</div>
<div style='font-size: 1.1rem; margin-bottom: 5px; margin-top: 10px;'>
<b>할인 후 금액 (공급가액) : {int(total_supply_price):,}</b> 원 
<span style='font-size: 0.9rem; opacity: 0.7;'> (개당 {int(per_unit_supply):,} 원)</span>
</div>
<div style='font-size: 1.1rem; margin-bottom: 10px;'>부가세 (10%) : <b>{int(vat):,}</b> 원</div>
<div style='border-top: 1px solid rgba(128,128,128,0.3); margin: 10px 0;'></div>

<!-- [Update] 디자인 변경: 배경 분홍색(#FFE6E6/Light Pink), 글자 빨간색(#FF4B4B) -->
<div style='background-color: #FFE6E6; border-radius: 8px; padding: 15px; margin-top: 15px; margin-bottom: 10px; text-align: center;'>
    <div style='font-size: 1.1rem; color: #FF4B4B; font-weight: 700; margin-bottom: 5px;'>고객 안내 금액 (VAT 포함)</div>
    <div style='font-size: 2.0rem; color: #FF4B4B; font-weight: 800;'>{int(grand_total):,} 원</div>
    <div style='font-size: 1.0rem; color: #D32F2F; font-weight: 600; margin-top: 5px;'>(개당 {int(per_unit_grand):,} 원)</div>
</div>
</div>
{download_html}
</div>""", unsafe_allow_html=True)

# =========================================================
# [BOTTOM] 프로젝트 관리 (화면 최하단)
# =========================================================
st.markdown("<div style='margin-top: 100px;'></div>", unsafe_allow_html=True)
st.markdown("---")
st.markdown("## 💾 프로젝트 관리 (수정용 저장)")
st.caption("현재 작업 내용을 JSON 파일로 저장하거나, 이전에 저장한 파일을 불러와 계속 수정할 수 있습니다.")

# 파일명 및 데이터 준비
current_state_bottom = {
    "customer_input": st.session_state.get("customer_input", ""),
    "selected_sheet": st.session_state.get("selected_sheet", ""),
    "base_qty": st.session_state.get("base_qty", 100),
    "is_express": st.session_state.get("is_express", False),
    "express_rate": st.session_state.get("express_rate", 1.2),
    "is_global_discount": st.session_state.get("is_global_discount", False),
    "global_discount_amt": st.session_state.get("global_discount_amt", 50000),
    "rows": st.session_state.get("rows", []),
    "next_id": st.session_state.get("next_id", 1)
}

today_str_bottom = datetime.now().strftime('%Y%m%d')
cust_name_bottom = clean_text(st.session_state.get("customer_input", "")).replace(" ", "_")
if not cust_name_bottom: cust_name_bottom = "Unknown"

# [다운로드용] 기본 파일명 (번호 없음)
json_filename_download = f"{today_str_bottom}_{cust_name_bottom}_estimate.json"
json_str_bottom = json.dumps(current_state_bottom, ensure_ascii=False, indent=2)

# 백업과 불러오기를 나란히 배치
col_backup, col_load = st.columns(2)

with col_backup:
    st.markdown("#### 💾 백업")
    if st.button("💾 로컬 백업 (backup 폴더)", use_container_width=True, key="save_local"):
        try:
            # 날짜별 폴더 생성 (backup/20260115/)
            backup_root = "backup"
            date_folder = os.path.join(backup_root, today_str_bottom)
            if not os.path.exists(date_folder): 
                os.makedirs(date_folder)
            
            # 같은 업체명 파일 찾아서 번호 매기기
            base_name = f"{cust_name_bottom}"
            counter = 1
            while True:
                filename = f"{base_name}_{counter}.json"
                file_path = os.path.join(date_folder, filename)
                if not os.path.exists(file_path):
                    break
                counter += 1
            
            # 파일 저장
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(json_str_bottom)
            
            st.success(f"✅ 저장 완료!\n📁 경로: `{date_folder}/{filename}`")
        except Exception as e:
            st.error(f"❌ 저장 실패: {e}")

with col_load:
    st.markdown("#### 📂 불러오기")
    uploaded_bottom = st.file_uploader("JSON 파일 선택", type=["json"], key="loader_bottom", label_visibility="collapsed")
    if uploaded_bottom is not None:
        try:
            data_loaded = json.load(uploaded_bottom)
            # Update all keys (에러 무시하고 rerun 강제)
            try:
                for k in ["customer_input", "selected_sheet", "base_qty", "is_express", "express_rate", "is_global_discount", "global_discount_amt"]:
                    if k in data_loaded: 
                        st.session_state[k] = data_loaded[k]
            except:
                pass  # 위젟 에러 무시
            
            # 핵심 데이터는 반드시 업데이트
            if "rows" in data_loaded: st.session_state["rows"] = data_loaded["rows"]
            if "next_id" in data_loaded: st.session_state["next_id"] = data_loaded["next_id"]
            
            # 무조건 rerun (파일 선택 즉시 적용)
            st.rerun()
        except Exception as e:
            st.error(f"❌ 파일 로드 실패: {e}")
